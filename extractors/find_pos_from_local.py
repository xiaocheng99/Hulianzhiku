import openpyxl
import pymysql
from openpyxl import load_workbook
import time
from fuzzywuzzy import process


# 数据库连接
def cnn():
    cnn = pymysql.connect(host='210.45.212.126',  # IP
                          user='hulianzhiku',  # 用户名
                          password='hulianzhiku',  # 密码
                          port=10010,  # 端口号
                          charset='utf8',
                          database='hulianzhiku')  # 注意是utf8不是utf-8

    return cnn


# 被诉主体点位查询
def find_pos(fileName, sheet, cnn):
    wb = load_workbook(fileName)
    ws = wb[sheet]
    result_str = ''
    result_list = []
    pos = ''
    EXCEL_col = 0
    for row in ws.rows:  # 获取每一行的数据
        EXCEL_col += 1
        count = 0
        for data in row:  # 获取每一行中单元格的数据
            if count == 1:
                main_parts = str(data.value).split('/')
                pos = compare_pos(main_parts, cnn)
                # 如果数据不为null 将点位写进点位表
                if pos != 'null':
                    print(pos)
                    ws['D'+str(EXCEL_col)] = pos
            count += 1
    print(EXCEL_col)
    wb.save(fileName)
    wb.close()


def compare_pos(main_parts, cnn):
    pos = compare_house(main_parts, cnn)
    if pos == 'null':
        pos = compare_company(main_parts, cnn)
    if pos == 'null':
        pos = compare_community(main_parts, cnn)
    return pos


def fuzzy_compare():
    users = [' ', ' 侯家桥路38号 ', ' N ', ' 金牛区侯家桥路38号', ' 花园 ', ' 侯家花园小区 ', ' 侯家桥路 ', ' 金牛区 ', ' 自来水公司 ', ' N']
    a = process.extractOne('金牛区侯家桥路38号侯家花园小区B区', users)
    print(a)


# 匹配公司点位
def compare_company(main_parts, cnn):
    pos = 'null'
    # 使用cursor()方法获取操作游标
    cursor = cnn.cursor()
    # 使用execute方法执行SQL语句
    cursor.execute("SELECT company FROM `company_PT`")
    results = cursor.fetchall()

    for res in results:
        for main_part in main_parts:
            main_part = main_part.strip()
            if main_part != 'N':
                # 进行匹配
                if main_part == res:
                    cursor.execute("SELECT PT FROM `company_PT` WHERE company = " + main_part)
                    pos = cursor.fetchone()[0]
                    return pos

    return pos


# 匹配社区点位
def compare_community(main_parts, cnn):
    pos = 'null'
    # 使用cursor()方法获取操作游标
    cursor = cnn.cursor()
    # 使用execute方法执行SQL语句
    cursor.execute("SELECT PT FROM `community_PT`")
    results = cursor.fetchall()
    for res in results:
        for main_part in main_parts:
            main_part = main_part.strip()
            if main_part != 'N':
                # 进行匹配
                if main_part == res:
                    cursor.execute("SELECT community FROM `company_PT` WHERE PT = " + main_part)
                    pos = cursor.fetchone()[0]
                    return pos

    return pos


def compare_house(main_parts, cnn):
    pos = 'null'
    # 使用cursor()方法获取操作游标
    cursor = cnn.cursor()
    # 使用execute方法执行SQL语句
    cursor.execute("SELECT house FROM `house_PT`")
    results = cursor.fetchall()
    for res in results:
        res = res[0].split('/')[1]
        # print(res)
        for main_part in main_parts:
            main_part = main_part.strip()
            if main_part != 'N':
                # 进行匹配
                if main_part == res:
                    cursor.execute("SELECT PT FROM `house_PT` WHERE house like " +'"'+"%"+ main_part +'"' )
                    pos = cursor.fetchone()[0]
                    return pos

    return pos


if __name__ == '__main__':
    start = time.time()

    cnn = cnn()
    find_pos('被诉主体地址_full.xlsx', '高新区法院', cnn)

    end = time.time()
    print('程序运行时间:' + str(end - start) + 's')
